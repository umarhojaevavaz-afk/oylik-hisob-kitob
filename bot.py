"""
Moliya boti — Excel faylga yozuvchi Telegram bot
O'zbek tilida | v3.0
Faqat BOT_TOKEN kerak, hech qanday Google yo'q!
"""

import os
import json
import logging
from datetime import datetime, date, timedelta
from pathlib import Path

from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, ReplyKeyboardMarkup
from telegram.ext import (
    Application, CommandHandler, MessageHandler, CallbackQueryHandler,
    ConversationHandler, filters, ContextTypes
)
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ─── Sozlamalar ───────────────────────────────────────────────────────────────
BOT_TOKEN = os.environ.get("BOT_TOKEN", "").strip()
EXCEL_DIR = Path(os.environ.get("EXCEL_DIR", "moliya_data"))
EXCEL_DIR.mkdir(exist_ok=True)
USERS_FILE = EXCEL_DIR / "users.json"

MONTH_NAMES = ["", "Yanvar", "Fevral", "Mart", "Aprel", "May", "Iyun",
               "Iyul", "Avgust", "Sentabr", "Oktabr", "Noyabr", "Dekabr"]

CHIQIM_KATEGORIYALAR = [
    "🍔 Oziq-ovqat", "🚗 Transport", "🏥 Tibbiy/Dorixona",
    "📚 Ta'lim", "🏠 Uy/Kommunal", "📱 Telefon/Internet",
    "👗 Kiyim-kechak", "💼 Biznes", "🎭 Ko'ngil ochar",
    "💸 P2P o'tkazma", "🏦 Bank to'lovi", "🛒 Savdo marketi",
    "👤 Shaxsiy to'lov", "🔧 Texnika", "📦 Boshqa",
]
KIRIM_KATEGORIYALAR = [
    "💰 Maosh/Daromad", "🏪 Biznes tushumi", "💳 Karta to'ldirish",
    "🏦 Bank o'tkazma", "🎁 Sovg'a", "💹 Cashback", "📦 Boshqa kirim",
]

(SELECTING_CATEGORY, ENTERING_AMOUNT, ENTERING_DESCRIPTION,
 SELECTING_DATE, CONFIRMING) = range(5)

logging.basicConfig(format="%(asctime)s %(levelname)s %(message)s", level=logging.INFO)
logger = logging.getLogger(__name__)


# ─── Excel funksiyalar ────────────────────────────────────────────────────────
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HDR_FILL = PatternFill("solid", start_color="1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)


def get_excel_path(year: int, month: int) -> Path:
    return EXCEL_DIR / f"{year}-{month:02d}-moliya.xlsx"


HEADERS = ["Sana", "Vaqt", "Tur", "Kategoriya", "Summa", "Valyuta", "Izoh"]
COL_WIDTHS = {"A": 12, "B": 8, "C": 12, "D": 22, "E": 18, "F": 10, "G": 30}


def _write_headers(ws):
    for col, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER
    for letter, width in COL_WIDTHS.items():
        ws.column_dimensions[letter].width = width
    ws.freeze_panes = "A2"


def get_or_create_workbook(year: int, month: int):
    path = get_excel_path(year, month)
    if path.exists():
        wb = openpyxl.load_workbook(path)
        ws = wb["Tranzaksiyalar"]
        # Eski fayllarda sarlavha 6 ta ustun edi (Valyuta yo'q, Izoh noto'g'ri
        # joyda). Ma'lumot qatorlariga tegmasdan sarlavhani yangilaymiz.
        if str(ws.cell(row=1, column=6).value or "") != "Valyuta":
            _write_headers(ws)
            wb.save(path)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Tranzaksiyalar"
        _write_headers(ws)
        wb.save(path)
    return wb, path


def add_transaction(data: dict) -> str:
    now = datetime.now()
    wb, path = get_or_create_workbook(now.year, now.month)
    ws = wb["Tranzaksiyalar"]

    is_income = "Kirim" in data["type"]
    fill = PatternFill("solid", start_color="E2EFDA" if is_income else "FCE4D6")
    font_color = "375623" if is_income else "843C0C"

    currency = data.get("currency", "UZS")
    # Summa doim raqam bo'lib yoziladi — Excelda yig'indi/filtr ishlashi uchun.
    row_data = [
        data.get("date", date.today().strftime("%d.%m.%Y")),
        now.strftime("%H:%M"),
        data["type"],
        data["category"].split(" ", 1)[-1],
        float(data["amount"]),
        currency,
        data["description"],
    ]
    next_row = ws.max_row + 1
    for col, val in enumerate(row_data, 1):
        cell = ws.cell(row=next_row, column=col, value=val)
        cell.border = BORDER
        cell.font = Font(name="Arial", size=10, color=font_color if col in [3, 4, 5] else "000000")
        cell.fill = fill if col == 3 else PatternFill()
        if col == 5:
            cell.number_format = "#,##0.00" if currency == "USD" else "#,##0"
            cell.alignment = Alignment(horizontal="right")
        elif col in [1, 2, 3, 6]:
            cell.alignment = Alignment(horizontal="center")

    wb.save(path)
    return str(path)


def _cell(row, idx):
    """Qator kalta bo'lsa ham xato bermaydi."""
    return row[idx] if len(row) > idx else None


def _parse_amount_currency(row):
    """row[4]=summa, row[5]=valyuta (eski fayllarda summa ichida "$" bo'lishi mumkin)"""
    raw = str(_cell(row, 4) if _cell(row, 4) is not None else "0")
    # Belgilarni tozalash: "$", "," va bo'sh joylar
    clean = raw.replace("$", "").replace(",", "").replace(" ", "").strip()
    try:
        amount = float(clean) if clean else 0.0
    except ValueError:
        amount = 0.0
    # Currency aniqlash
    currency = "UZS"
    cur_cell = _cell(row, 5)
    if "$" in raw:
        currency = "USD"
    elif cur_cell is not None and str(cur_cell).strip().upper() in ("USD", "UZS"):
        currency = str(cur_cell).strip().upper()
    return amount, currency


def _fmt(amount: float, currency: str) -> str:
    return f"{amount:,.2f} $" if currency == "USD" else f"{amount:,.0f} UZS"


def _row_description(row) -> str:
    """Izoh: yangi formatda 7-ustun, eski formatda 6-ustunda bo'lishi mumkin."""
    desc = _cell(row, 6)
    if desc:
        return str(desc)
    old = _cell(row, 5)
    if old and str(old).strip().upper() not in ("USD", "UZS"):
        return str(old)
    return ""


def get_monthly_summary(year: int, month: int) -> dict | None:
    path = get_excel_path(year, month)
    if not path.exists():
        return None
    wb = openpyxl.load_workbook(path)
    ws = wb["Tranzaksiyalar"]
    total_in_uzs, total_out_uzs = 0.0, 0.0
    total_in_usd, total_out_usd = 0.0, 0.0
    # Kategoriyalar valyuta bo'yicha alohida yig'iladi — UZS va USD ni
    # qo'shib yuborish summani ma'nosiz qilib qo'yadi.
    cats_out = {"UZS": {}, "USD": {}}
    cats_in = {"UZS": {}, "USD": {}}
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not _cell(row, 0):
            continue
        tur = str(_cell(row, 2) or "")
        kat = str(_cell(row, 3) or "Boshqa")
        amount, currency = _parse_amount_currency(row)
        if amount == 0:
            continue
        count += 1
        if "Kirim" in tur:
            if currency == "USD":
                total_in_usd += amount
            else:
                total_in_uzs += amount
            bucket = cats_in[currency]
        else:
            if currency == "USD":
                total_out_usd += amount
            else:
                total_out_uzs += amount
            bucket = cats_out[currency]
        bucket[kat] = bucket.get(kat, 0.0) + amount
    return {"total_in": total_in_uzs, "total_out": total_out_uzs,
            "total_in_usd": total_in_usd, "total_out_usd": total_out_usd,
            "net": total_in_uzs - total_out_uzs,
            "net_usd": total_in_usd - total_out_usd,
            "cats_in": cats_in, "cats_out": cats_out, "count": count}


def _top_cats(cats: dict, limit: int, totals: dict | None = None) -> str:
    """Valyuta bo'yicha eng katta kategoriyalar ro'yxati."""
    text = ""
    for currency in ("UZS", "USD"):
        items = sorted(cats.get(currency, {}).items(), key=lambda x: -x[1])[:limit]
        for kat, val in items:
            line = f"  • {kat}: {_fmt(val, currency)}"
            if totals and totals.get(currency):
                line += f" ({val / totals[currency] * 100:.0f}%)"
            text += line + "\n"
    return text


def get_today_records() -> list:
    now = datetime.now()
    path = get_excel_path(now.year, now.month)
    if not path.exists():
        return []
    today_str = date.today().strftime("%d.%m.%Y")
    wb = openpyxl.load_workbook(path)
    ws = wb["Tranzaksiyalar"]
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == today_str:
            records.append(row)
    return records


# ─── Telegram handlers ────────────────────────────────────────────────────────
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    keyboard = [
        ["➕ Xarajat qo'shish", "💰 Daromad qo'shish"],
        ["📊 Bu oy statistikasi", "📅 O'tgan oy hisoboti"],
        ["📋 Bugungi yozuvlar", "❓ Yordam"],
    ]
    await update.message.reply_text(
        "👋 *Salom! Moliya botiga xush kelibsiz!*\n\n"
        "Kunlik xarajat va daromadlaringizni yozing.\n"
        "Barchasi Excel fayliga avtomatik saqlanadi. 📊\n\n"
        "📁 Fayllar: `moliya_data/` papkasida",
        parse_mode="Markdown",
        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True),
    )


async def handle_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    if text == "➕ Xarajat qo'shish":
        context.user_data["transaction_type"] = "🔴 Chiqim"
        return await _show_categories(update, CHIQIM_KATEGORIYALAR, "💸 Xarajat kategoriyasini tanlang:")
    elif text == "💰 Daromad qo'shish":
        context.user_data["transaction_type"] = "💚 Kirim"
        return await _show_categories(update, KIRIM_KATEGORIYALAR, "💰 Daromad kategoriyasini tanlang:")
    elif text == "📊 Bu oy statistikasi":
        await _current_stats(update)
    elif text == "📅 O'tgan oy hisoboti":
        await _last_month(update)
    elif text == "📋 Bugungi yozuvlar":
        await _today(update)
    elif text == "❓ Yordam":
        await _help(update)
    return ConversationHandler.END


async def _show_categories(update, categories, prompt):
    cols = 2
    keyboard = []
    for i in range(0, len(categories), cols):
        row = [InlineKeyboardButton(c, callback_data=f"cat:{c}") for c in categories[i:i+cols]]
        keyboard.append(row)
    keyboard.append([InlineKeyboardButton("❌ Bekor", callback_data="cancel")])
    await update.message.reply_text(prompt, reply_markup=InlineKeyboardMarkup(keyboard))
    return SELECTING_CATEGORY


async def category_selected(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if query.data == "cancel":
        await query.edit_message_text("❌ Bekor qilindi.")
        return ConversationHandler.END
    context.user_data["category"] = query.data.replace("cat:", "")
    await query.edit_message_text(
        f"✅ Kategoriya: *{context.user_data['category']}*\n\n"
        f"💵 Summani kiriting (UZS):\n_Masalan: 50000_",
        parse_mode="Markdown"
    )
    return ENTERING_AMOUNT


async def amount_entered(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip().replace(",", "").replace(" ", "")
    # Valyuta turini aniqlash
    currency = "UZS"
    if "$" in text or "usd" in text.lower() or "dollar" in text.lower():
        currency = "USD"
    # Belgilarni olib tashlash
    text_clean = text.replace("$", "").replace("usd", "").replace("USD", "").replace("uzs", "").replace("UZS", "").strip()
    try:
        amount = float(text_clean)
        if amount <= 0:
            raise ValueError
    except ValueError:
        await update.message.reply_text(
            "❌ Faqat raqam kiriting:\n_Masalan: 50000 yoki 540$ yoki 100 USD_",
            parse_mode="Markdown"
        )
        return ENTERING_AMOUNT
    context.user_data["amount"] = amount
    context.user_data["currency"] = currency
    if currency == "USD":
        display = f"{amount:,.2f} $"
    else:
        display = f"{int(amount):,} UZS"
    await update.message.reply_text(
        f"✅ Summa: *{display}*\n\n📝 Izoh kiriting:\n_Masalan: Tushlik, Maosh, Dorixona_",
        parse_mode="Markdown"
    )
    return ENTERING_DESCRIPTION


async def description_entered(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["description"] = update.message.text.strip()
    context.user_data["date"] = date.today().strftime("%d.%m.%Y")
    keyboard = [
        [InlineKeyboardButton(f"📅 Bugun ({date.today().strftime('%d.%m')})", callback_data="date:today"),
         InlineKeyboardButton(f"📅 Kecha ({(date.today()-timedelta(1)).strftime('%d.%m')})", callback_data="date:yesterday")],
        [InlineKeyboardButton("✏️ Boshqa sana kiriting", callback_data="date:custom")],
    ]
    await update.message.reply_text("📅 Sana tanlang:", reply_markup=InlineKeyboardMarkup(keyboard))
    return SELECTING_DATE


async def date_selected(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if query.data == "date:today":
        context.user_data["date"] = date.today().strftime("%d.%m.%Y")
    elif query.data == "date:yesterday":
        context.user_data["date"] = (date.today() - timedelta(1)).strftime("%d.%m.%Y")
    elif query.data == "date:custom":
        await query.edit_message_text("📅 Sanani kiriting (KK.OO.YYYY):\n_Masalan: 15.05.2026_",
                                       parse_mode="Markdown")
        context.user_data["waiting_custom_date"] = True
        return SELECTING_DATE
    return await _show_confirm_query(query, context)


async def custom_date_msg(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not context.user_data.get("waiting_custom_date"):
        return SELECTING_DATE
    try:
        d = datetime.strptime(update.message.text.strip(), "%d.%m.%Y")
        context.user_data["date"] = d.strftime("%d.%m.%Y")
        context.user_data["waiting_custom_date"] = False
    except ValueError:
        await update.message.reply_text("❌ Format noto'g'ri. KK.OO.YYYY:\n_Masalan: 15.05.2026_",
                                         parse_mode="Markdown")
        return SELECTING_DATE
    d = context.user_data
    keyboard = [[InlineKeyboardButton("✅ Saqlash", callback_data="confirm:yes"),
                 InlineKeyboardButton("❌ Bekor", callback_data="confirm:no")]]
    await update.message.reply_text(
        _confirm_text(d), parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )
    return CONFIRMING


def _confirm_text(d):
    currency = d.get("currency", "UZS")
    amount_str = f"{d['amount']:,.2f} $" if currency == "USD" else f"{int(d['amount']):,} UZS"
    return (f"📋 *Tasdiqlang:*\n\n"
            f"{'Tur:':<14} {d['transaction_type']}\n"
            f"{'Kategoriya:':<14} {d['category']}\n"
            f"{'Summa:':<14} *{amount_str}*\n"
            f"{'Izoh:':<14} {d['description']}\n"
            f"{'Sana:':<14} {d['date']}")


async def _show_confirm_query(query, context):
    d = context.user_data
    keyboard = [[InlineKeyboardButton("✅ Saqlash", callback_data="confirm:yes"),
                 InlineKeyboardButton("❌ Bekor", callback_data="confirm:no")]]
    await query.edit_message_text(_confirm_text(d), parse_mode="Markdown",
                                   reply_markup=InlineKeyboardMarkup(keyboard))
    return CONFIRMING


async def confirm_transaction(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if query.data == "confirm:no":
        await query.edit_message_text("❌ Bekor qilindi.")
        return ConversationHandler.END
    d = context.user_data
    try:
        file_path = add_transaction({
            "type": d["transaction_type"],
            "category": d["category"],
            "amount": d["amount"],
            "currency": d.get("currency", "UZS"),
            "description": d["description"],
            "date": d["date"],
        })
        emoji = "💚" if "Kirim" in d["transaction_type"] else "🔴"
        file_name = Path(file_path).name
        currency = d.get("currency", "UZS")
        amount_str = f"{d['amount']:,.2f} $" if currency == "USD" else f"{int(d['amount']):,} UZS"
        await query.edit_message_text(
            f"✅ *Saqlandi!*\n\n"
            f"{emoji} *{amount_str}* — {d['description']}\n"
            f"📁 {d['category']}\n"
            f"📅 {d['date']}\n\n"
            f"💾 Fayl: `{file_name}`",
            parse_mode="Markdown"
        )
    except Exception as e:
        logger.error(f"Excel xatosi: {e}")
        await query.edit_message_text(f"❌ Xato: {str(e)}")
    context.user_data.clear()
    return ConversationHandler.END


async def _current_stats(update):
    now = datetime.now()
    msg = await update.message.reply_text("⏳ Hisoblanmoqda...")
    s = get_monthly_summary(now.year, now.month)
    if not s or s["count"] == 0:
        await msg.edit_text("📭 Bu oy hali yozuv yo'q.\n\nXarajat yoki daromad qo'shing!")
        return
    net_e = "✅" if s["net"] >= 0 else "⚠️"
    text = f"📊 *{now.year}-yil {MONTH_NAMES[now.month]} oyi*\n\n"
    # UZS
    if s["total_in"] or s["total_out"]:
        text += (f"💚 Kirim (UZS):  *{s['total_in']:,.0f} UZS*\n"
                 f"🔴 Chiqim (UZS): *{s['total_out']:,.0f} UZS*\n"
                 f"{net_e} Balans (UZS): *{s['net']:,.0f} UZS*\n")
    # USD
    if s["total_in_usd"] or s["total_out_usd"]:
        net_usd_e = "✅" if s["net_usd"] >= 0 else "⚠️"
        text += (f"\n💚 Kirim (USD):  *{s['total_in_usd']:,.2f} $*\n"
                 f"🔴 Chiqim (USD): *{s['total_out_usd']:,.2f} $*\n"
                 f"{net_usd_e} Balans (USD): *{s['net_usd']:,.2f} $*\n")
    text += f"\n📝 Jami yozuvlar: {s['count']} ta\n"
    top = _top_cats(s["cats_out"], 5)
    if top:
        text += "\n🔴 *Eng katta chiqimlar:*\n" + top
    await msg.edit_text(text, parse_mode="Markdown")


async def _last_month(update):
    now = datetime.now()
    year, month = (now.year - 1, 12) if now.month == 1 else (now.year, now.month - 1)
    msg = await update.message.reply_text("⏳ Yuklanmoqda...")
    s = get_monthly_summary(year, month)
    if not s or s["count"] == 0:
        await msg.edit_text(f"📭 {MONTH_NAMES[month]} oyi uchun ma'lumot yo'q.")
        return
    net_e = "✅" if s["net"] >= 0 else "⚠️"
    text = f"📅 *{year}-yil {MONTH_NAMES[month]} oyi*\n\n"
    if s["total_in"] or s["total_out"]:
        text += (f"💚 Jami kirim:  *{s['total_in']:,.0f} UZS*\n"
                 f"🔴 Jami chiqim: *{s['total_out']:,.0f} UZS*\n"
                 f"{net_e} Sof balans:  *{s['net']:,.0f} UZS*\n")
    if s["total_in_usd"] or s["total_out_usd"]:
        net_usd_e = "✅" if s["net_usd"] >= 0 else "⚠️"
        text += (f"\n💚 Jami kirim:  *{s['total_in_usd']:,.2f} $*\n"
                 f"🔴 Jami chiqim: *{s['total_out_usd']:,.2f} $*\n"
                 f"{net_usd_e} Sof balans:  *{s['net_usd']:,.2f} $*\n")
    text += f"\n📝 Jami: {s['count']} ta\n"
    totals = {"UZS": s["total_out"], "USD": s["total_out_usd"]}
    top = _top_cats(s["cats_out"], 6, totals)
    if top:
        text += "\n🔴 *Chiqimlar:*\n" + top
    await msg.edit_text(text, parse_mode="Markdown")


async def _today(update):
    records = get_today_records()
    today_str = date.today().strftime("%d.%m.%Y")
    if not records:
        await update.message.reply_text(f"📭 {today_str} — bugun hali yozuv yo'q.")
        return
    text = f"📋 *{today_str} — Bugungi yozuvlar*\n\n"
    tin_uzs = tout_uzs = tin_usd = tout_usd = 0.0
    for r in records:
        is_income = "Kirim" in str(_cell(r, 2) or "")
        e = "💚" if is_income else "🔴"
        amount, currency = _parse_amount_currency(r)
        desc = _row_description(r)
        if currency == "USD":
            if is_income:
                tin_usd += amount
            else:
                tout_usd += amount
        else:
            if is_income:
                tin_uzs += amount
            else:
                tout_uzs += amount
        text += f"{e} {_cell(r, 3)} — *{_fmt(amount, currency)}* — _{desc}_\n"
    text += "\n"
    if tin_uzs or tout_uzs:
        text += f"💚 Kirim: {tin_uzs:,.0f} | 🔴 Chiqim: {tout_uzs:,.0f} UZS\n"
    if tin_usd or tout_usd:
        text += f"💚 Kirim: {tin_usd:,.2f} | 🔴 Chiqim: {tout_usd:,.2f} $"
    await update.message.reply_text(text, parse_mode="Markdown")


async def _help(update):
    await update.message.reply_text(
        "❓ *Yordam*\n\n"
        "➕ *Xarajat/Daromad qo'shish:*\n"
        "Kategoriya → Summa → Izoh → Sana → Tasdiqlash\n\n"
        "📊 *Ko'rsatkichlar:*\n"
        "• Bu oy statistikasi\n"
        "• O'tgan oy hisoboti\n"
        "• Bugungi yozuvlar\n\n"
        "💾 *Fayllar:* `moliya_data/` papkasida\n"
        "Har oy alohida Excel fayl yaratiladi.\n"
        "Masalan: `2026-05-moliya.xlsx`",
        parse_mode="Markdown"
    )


async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data.clear()
    await update.message.reply_text("❌ Bekor qilindi.")
    return ConversationHandler.END


async def monthly_report_job(context: ContextTypes.DEFAULT_TYPE):
    now = datetime.now()
    year, month = (now.year - 1, 12) if now.month == 1 else (now.year, now.month - 1)
    s = get_monthly_summary(year, month)
    if not s or s["count"] == 0:
        return
    net_e = "✅" if s["net"] >= 0 else "⚠️"
    for uid in load_users():
        try:
            await context.bot.send_message(
                chat_id=uid,
                text=(f"🗓 *{year}-yil {MONTH_NAMES[month]} oyi tugadi!*\n\n"
                      f"💚 Kirim: *{s['total_in']:,.0f} UZS*\n"
                      f"🔴 Chiqim: *{s['total_out']:,.0f} UZS*\n"
                      f"{net_e} Balans: *{s['net']:,.0f} UZS*\n\n"
                      f"Yangi oy muborak! 💪"),
                parse_mode="Markdown"
            )
        except Exception as e:
            logger.error(f"Hisobot xatosi: {e}")


async def reset_month(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Joriy oy Excel faylini tozalash — avval tasdiq so'raladi"""
    now = datetime.now()
    path = get_excel_path(now.year, now.month)
    if not path.exists():
        await update.message.reply_text("📭 Bu oyda hali yozuv yo'q edi.")
        return
    s = get_monthly_summary(now.year, now.month)
    count = s["count"] if s else 0
    keyboard = [[InlineKeyboardButton("🗑 Ha, o'chirilsin", callback_data="reset:yes"),
                 InlineKeyboardButton("❌ Yo'q", callback_data="reset:no")]]
    await update.message.reply_text(
        f"⚠️ *Diqqat!*\n\n{now.year}-yil {MONTH_NAMES[now.month]} oyidagi "
        f"*{count} ta yozuv* butunlay o'chiriladi.\nBuni orqaga qaytarib bo'lmaydi.\n\n"
        f"Rostdan o'chirilsinmi?",
        parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup(keyboard),
    )


async def reset_confirmed(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if query.data == "reset:no":
        await query.edit_message_text("✅ Bekor qilindi — hech narsa o'chirilmadi.")
        return
    now = datetime.now()
    path = get_excel_path(now.year, now.month)
    if not path.exists():
        await query.edit_message_text("📭 Bu oyda hali yozuv yo'q edi.")
        return
    # O'chirishdan oldin zaxira nusxa qoldiramiz.
    backup = path.with_suffix(f".backup-{now.strftime('%Y%m%d-%H%M%S')}.xlsx")
    path.rename(backup)
    await query.edit_message_text(
        f"🗑 *Joriy oy ma'lumotlari tozalandi!*\n\n"
        f"💾 Zaxira nusxa: `{backup.name}`\n"
        f"Endi yangi yozuvlar qo'shishingiz mumkin.",
        parse_mode="Markdown"
    )


def load_users() -> set:
    try:
        return set(json.loads(USERS_FILE.read_text()))
    except (OSError, ValueError):
        return set()


def save_user(user_id: int) -> None:
    users = load_users()
    if user_id in users:
        return
    users.add(user_id)
    try:
        USERS_FILE.write_text(json.dumps(sorted(users)))
    except OSError as e:
        logger.error(f"users.json saqlanmadi: {e}")


async def track_user(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Foydalanuvchini diskka yozamiz — bot qayta ishga tushsa ham yo'qolmaydi."""
    if update.effective_user:
        save_user(update.effective_user.id)


def main():
    if not BOT_TOKEN:
        raise SystemExit(
            "❌ BOT_TOKEN topilmadi.\n"
            "Tokenni muhit o'zgaruvchisi orqali bering:\n"
            "  export BOT_TOKEN='...'   (yoki hosting panelidagi Environment Variables)"
        )
    app = Application.builder().token(BOT_TOKEN).build()

    conv = ConversationHandler(
        entry_points=[MessageHandler(
            filters.Regex("^(➕ Xarajat qo'shish|💰 Daromad qo'shish)$"), handle_menu)],
        states={
            SELECTING_CATEGORY: [CallbackQueryHandler(category_selected)],
            ENTERING_AMOUNT: [MessageHandler(filters.TEXT & ~filters.COMMAND, amount_entered)],
            ENTERING_DESCRIPTION: [MessageHandler(filters.TEXT & ~filters.COMMAND, description_entered)],
            SELECTING_DATE: [
                CallbackQueryHandler(date_selected, pattern="^date:"),
                MessageHandler(filters.TEXT & ~filters.COMMAND, custom_date_msg),
            ],
            CONFIRMING: [CallbackQueryHandler(confirm_transaction, pattern="^confirm:")],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
    )

    app.add_handler(MessageHandler(filters.ALL, track_user), group=-1)
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("reset", reset_month))
    app.add_handler(CallbackQueryHandler(reset_confirmed, pattern="^reset:"))
    app.add_handler(conv)
    app.add_handler(MessageHandler(
        filters.Regex("^(📊 Bu oy statistikasi|📅 O'tgan oy hisoboti|📋 Bugungi yozuvlar|❓ Yordam)$"),
        handle_menu
    ))

    # Har oy 1-si soat 09:00 da hisobot
    from datetime import time as dtime
    app.job_queue.run_monthly(
        monthly_report_job,
        when=dtime(9, 0, 0),
        day=1
    )

    logger.info("✅ Moliya boti ishga tushdi!")
    logger.info(f"📁 Excel fayllar: {EXCEL_DIR.absolute()}")
    app.run_polling(allowed_updates=Update.ALL_TYPES)


if __name__ == "__main__":
    main()
